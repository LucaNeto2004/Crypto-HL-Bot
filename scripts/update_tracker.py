"""
Excel Tracker Updater — runs hourly via cron (6am-10pm).
Rebuilds ALL daily rows from paper_state.json so cumulative P&L is always correct.
"""
import json
import os
import sys
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

TRACKER_PATH = os.path.expanduser("~/Desktop/Bot_Investment_Tracker.xlsx")
STATE_PATH = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "data", "paper_state.json")

green_font = Font(color='22C55E', bold=True)
red_font = Font(color='EF4444', bold=True)
green_bg = PatternFill(start_color='DCFCE7', end_color='DCFCE7', fill_type='solid')
red_bg = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid')
orange_bg = PatternFill(start_color='FEF3C7', end_color='FEF3C7', fill_type='solid')
orange_font = Font(color='D97706', bold=True)
live_bg = PatternFill(start_color='DCFCE7', end_color='DCFCE7', fill_type='solid')
live_font = Font(color='16A34A', bold=True)


def run():
    if not os.path.exists(TRACKER_PATH):
        print(f"Tracker not found: {TRACKER_PATH}")
        sys.exit(1)

    with open(STATE_PATH) as f:
        d = json.load(f)

    MAKER_FEE_PCT = 0.006 / 100  # 0.006% blended fee (limit entries + market exits)

    # Build daily stats from ALL trade history
    days = {}
    for t in d.get("trade_history", []):
        day = t["timestamp"][:10]
        pnl = t.get("pnl")
        volume = (t.get("size", 0) or 0) * (t.get("price", 0) or 0)
        if day not in days:
            days[day] = {"pnl": 0, "wins": 0, "losses": 0, "trades": 0, "volume": 0}
        days[day]["volume"] += volume
        if pnl is not None:
            days[day]["pnl"] += pnl
            days[day]["trades"] += 1
            if pnl > 0:
                days[day]["wins"] += 1
            else:
                days[day]["losses"] += 1

    wb = openpyxl.load_workbook(TRACKER_PATH)
    ws = wb["Daily P&L"]

    # Ensure headers are set (row 3)
    headers = ['Date', 'Day', 'Mode', 'Trades', 'Wins', 'Losses', 'Win Rate',
               'Gross P&L', 'Gross Cumul. P&L', 'Est. Fees', 'Net P&L', 'Cumul. P&L', 'Cumul. %', 'Balance', 'Notes']
    # Column widths for readability
    col_widths = {
        'A': 12, 'B': 10, 'C': 8, 'D': 8, 'E': 7, 'F': 7, 'G': 9,
        'H': 12, 'I': 16, 'J': 10, 'K': 12, 'L': 13, 'M': 10, 'N': 13, 'O': 25
    }
    for col, w in col_widths.items():
        ws.column_dimensions[col].width = w

    thin_border = Border(
        left=Side(style='thin', color='D1D5DB'),
        right=Side(style='thin', color='D1D5DB'),
        top=Side(style='thin', color='D1D5DB'),
        bottom=Side(style='thin', color='D1D5DB'),
    )
    center = Alignment(horizontal='center', vertical='center')
    center_wrap = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left = Alignment(horizontal='left', vertical='center')

    for i, h in enumerate(headers, 1):
        hc = ws.cell(row=3, column=i, value=h)
        hc.font = Font(bold=True, color='FFFFFF', size=10)
        hc.fill = PatternFill(start_color='0D1B2A', end_color='0D1B2A', fill_type='solid')
        hc.alignment = center_wrap
        hc.border = thin_border

    # Set row height for header
    ws.row_dimensions[3].height = 30

    cumul = 0
    gross_cumul = 0
    r = 4
    for day in sorted(days.keys()):
        s = days[day]
        wr = s["wins"] / s["trades"] if s["trades"] > 0 else 0
        gross = round(s["pnl"], 2)
        fees = round(s["volume"] * MAKER_FEE_PCT, 2)
        net = round(gross - fees, 2)
        cumul += net
        gross_cumul += gross

        ws.row_dimensions[r].height = 22

        dc = ws.cell(row=r, column=1, value=day)
        dc.alignment = center
        dc.font = Font(size=10)
        ws.cell(row=r, column=2, value="").alignment = center
        mode_cell = ws.cell(row=r, column=3, value="Paper")
        mode_cell.fill = orange_bg
        mode_cell.font = orange_font
        mode_cell.alignment = center
        ws.cell(row=r, column=4, value=s["trades"]).alignment = center
        wc = ws.cell(row=r, column=5, value=s["wins"])
        wc.font = Font(color='22C55E')
        wc.alignment = center
        lc = ws.cell(row=r, column=6, value=s["losses"])
        lc.font = Font(color='EF4444')
        lc.alignment = center
        wrc = ws.cell(row=r, column=7, value=round(wr, 3))
        wrc.number_format = '0.00%'
        wrc.alignment = center

        gc = ws.cell(row=r, column=8, value=gross)
        gc.number_format = '$#,##0.00'
        gc.font = green_font if gross >= 0 else red_font
        gc.fill = green_bg if gross >= 0 else red_bg
        gc.alignment = center

        # Gross Cumulative P&L (no fees)
        gcm = ws.cell(row=r, column=9, value=round(gross_cumul, 2))
        gcm.number_format = '$#,##0.00'
        gcm.font = green_font if gross_cumul >= 0 else red_font
        gcm.fill = green_bg if gross_cumul >= 0 else red_bg
        gcm.alignment = center

        fc = ws.cell(row=r, column=10, value=fees)
        fc.number_format = '$#,##0.00'
        fc.font = Font(size=10)
        fc.alignment = center

        # Net P&L (gross - fees)
        nc = ws.cell(row=r, column=11, value=net)
        nc.number_format = '$#,##0.00'
        nc.font = green_font if net >= 0 else red_font
        nc.fill = green_bg if net >= 0 else red_bg
        nc.alignment = center

        # Net Cumulative
        cc = ws.cell(row=r, column=12, value=round(cumul, 2))
        cc.number_format = '$#,##0.00'
        cc.font = green_font if cumul >= 0 else red_font
        cc.fill = green_bg if cumul >= 0 else red_bg
        cc.alignment = center

        # Cumulative %
        cumul_pct = cumul / 10000
        cp = ws.cell(row=r, column=13, value=round(cumul_pct, 4))
        cp.number_format = '0.00%'
        cp.font = green_font if cumul_pct >= 0 else red_font
        cp.fill = green_bg if cumul_pct >= 0 else red_bg
        cp.alignment = center

        # Balance
        bc = ws.cell(row=r, column=14, value=round(10000 + cumul, 2))
        bc.number_format = '$#,##0.00'
        bc.font = Font(bold=True, size=10)
        bc.alignment = center

        # Notes
        ws.cell(row=r, column=15, value="").alignment = left

        # Apply borders to all columns
        for col in range(1, 16):
            ws.cell(row=r, column=col).border = thin_border

        r += 1

    # Extend borders to row 50 for clean table look
    for empty_r in range(r, 51):
        ws.row_dimensions[empty_r].height = 22
        for col in range(1, 16):
            cell = ws.cell(row=empty_r, column=col)
            cell.border = thin_border
            cell.alignment = center if col != 15 else left

    # === RECOUP TRACKER (right side of Daily P&L sheet) ===
    # Setup costs
    SETUP_COSTS = {
        "Mac Mini M2": 662.16,
        "Monitor": 140.54,
        "Keyboard": 162.16,
        "Laptop Stand": 32.43,
        "HL Deposit": 102.00,
    }
    MONTHLY_COSTS = {"Claude Max": 100.00}
    MONTHS_RUNNING = max(1, (len(days)))  # rough months from days of data

    # Column R onwards (col 18+)
    recap_start_col = 18
    recap_header_fill = PatternFill(start_color='1E3A5F', end_color='1E3A5F', fill_type='solid')
    recap_header_font = Font(bold=True, color='FFFFFF', size=11)

    # Title
    tc = ws.cell(row=3, column=recap_start_col, value="RECOUP TRACKER")
    tc.font = recap_header_font
    tc.fill = recap_header_fill
    tc.alignment = center
    ws.merge_cells(start_row=3, start_column=recap_start_col, end_row=3, end_column=recap_start_col + 1)
    ws.column_dimensions['R'].width = 20
    ws.column_dimensions['S'].width = 14

    # Headers
    for i, h in enumerate(["Item", "Amount"]):
        hc = ws.cell(row=4, column=recap_start_col + i, value=h)
        hc.font = Font(bold=True, color='FFFFFF', size=10)
        hc.fill = PatternFill(start_color='0D1B2A', end_color='0D1B2A', fill_type='solid')
        hc.alignment = center
        hc.border = thin_border

    # One-time costs
    row = 5
    total_setup = 0
    for item, cost in SETUP_COSTS.items():
        ws.cell(row=row, column=recap_start_col, value=item).alignment = left
        c = ws.cell(row=row, column=recap_start_col + 1, value=cost)
        c.number_format = '$#,##0.00'
        c.alignment = center
        for col in range(recap_start_col, recap_start_col + 2):
            ws.cell(row=row, column=col).border = thin_border
        total_setup += cost
        row += 1

    # Monthly costs
    for item, cost in MONTHLY_COSTS.items():
        ws.cell(row=row, column=recap_start_col, value=f"{item} (monthly)").alignment = left
        c = ws.cell(row=row, column=recap_start_col + 1, value=cost)
        c.number_format = '$#,##0.00'
        c.font = Font(color='D97706')
        c.alignment = center
        for col in range(recap_start_col, recap_start_col + 2):
            ws.cell(row=row, column=col).border = thin_border
        row += 1

    # Total invested
    total_monthly = sum(MONTHLY_COSTS.values()) * MONTHS_RUNNING
    total_invested = total_setup + total_monthly
    row += 1
    ws.cell(row=row, column=recap_start_col, value="Total One-Time").font = Font(bold=True)
    ws.cell(row=row, column=recap_start_col, value="Total One-Time").alignment = left
    c = ws.cell(row=row, column=recap_start_col + 1, value=total_setup)
    c.number_format = '$#,##0.00'
    c.font = Font(bold=True)
    c.alignment = center
    for col in range(recap_start_col, recap_start_col + 2):
        ws.cell(row=row, column=col).border = thin_border
    row += 1

    ws.cell(row=row, column=recap_start_col, value=f"Total Monthly ({MONTHS_RUNNING}mo)").font = Font(bold=True)
    ws.cell(row=row, column=recap_start_col).alignment = left
    c = ws.cell(row=row, column=recap_start_col + 1, value=total_monthly)
    c.number_format = '$#,##0.00'
    c.font = Font(bold=True)
    c.alignment = center
    for col in range(recap_start_col, recap_start_col + 2):
        ws.cell(row=row, column=col).border = thin_border
    row += 1

    ws.cell(row=row, column=recap_start_col, value="TOTAL INVESTED").font = Font(bold=True, size=11)
    ws.cell(row=row, column=recap_start_col).alignment = left
    c = ws.cell(row=row, column=recap_start_col + 1, value=total_invested)
    c.number_format = '$#,##0.00'
    c.font = Font(bold=True, size=11, color='EF4444')
    c.alignment = center
    for col in range(recap_start_col, recap_start_col + 2):
        ws.cell(row=row, column=col).border = thin_border
    row += 2

    # Net P&L vs Invested — only count LIVE profits (paper doesn't recoup real costs)
    # TODO: switch to live PnL when bot goes live
    LIVE_MODE = False  # Set True when trading live
    live_pnl = round(cumul, 2) if LIVE_MODE else 0.0

    ws.cell(row=row, column=recap_start_col, value="Live Net P&L").alignment = left
    c = ws.cell(row=row, column=recap_start_col + 1, value=live_pnl)
    c.number_format = '$#,##0.00'
    c.font = green_font if live_pnl >= 0 else red_font
    c.fill = green_bg if live_pnl > 0 else red_bg if live_pnl < 0 else orange_bg
    c.alignment = center
    for col in range(recap_start_col, recap_start_col + 2):
        ws.cell(row=row, column=col).border = thin_border
    row += 1

    remaining = total_invested - max(live_pnl, 0)
    ws.cell(row=row, column=recap_start_col, value="Remaining to Recoup").font = Font(bold=True)
    ws.cell(row=row, column=recap_start_col).alignment = left
    c = ws.cell(row=row, column=recap_start_col + 1, value=round(remaining, 2))
    c.number_format = '$#,##0.00'
    c.font = Font(bold=True, size=11, color='EF4444')
    c.alignment = center
    for col in range(recap_start_col, recap_start_col + 2):
        ws.cell(row=row, column=col).border = thin_border
    row += 1

    recoup_pct = max(live_pnl, 0) / total_invested if total_invested > 0 else 0
    ws.cell(row=row, column=recap_start_col, value="Recoup Progress").font = Font(bold=True)
    ws.cell(row=row, column=recap_start_col).alignment = left
    c = ws.cell(row=row, column=recap_start_col + 1, value=round(recoup_pct, 4))
    c.number_format = '0.0%'
    c.font = green_font if recoup_pct > 0 else red_font
    c.alignment = center
    for col in range(recap_start_col, recap_start_col + 2):
        ws.cell(row=row, column=col).border = thin_border
    row += 1

    # Mode indicator
    mode_label = "LIVE" if LIVE_MODE else "PAPER MODE"
    mode_color = '16A34A' if LIVE_MODE else 'D97706'
    ws.cell(row=row, column=recap_start_col, value="Status").alignment = left
    mc = ws.cell(row=row, column=recap_start_col + 1, value=mode_label)
    mc.font = Font(bold=True, color=mode_color)
    mc.alignment = center
    for col in range(recap_start_col, recap_start_col + 2):
        ws.cell(row=row, column=col).border = thin_border

    wb.save(TRACKER_PATH)
    print(f"Updated {len(days)} days | Cumul: ${cumul:+.2f} | Balance: ${10000+cumul:.2f}")


if __name__ == "__main__":
    run()
